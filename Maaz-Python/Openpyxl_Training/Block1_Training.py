import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from os import system

if system("dir C:\\Users\\Admin\\AppData\\Local\\Temp\\Block1_Training.xlsx >nul 2>&1") == 0:
    system("del C:\\Users\\Admin\\AppData\\Local\\Temp\\Block1_Training.xlsx  >nul 2>&1")
    print("File existed but I deleted it")

# ---------------------------
# USER STARTING WEIGHTS (adjusted for 17kg barbell)
# ---------------------------
starting_weights = {
    "Squat (Strength)": "117kg x 3",
    "Bench (Strength)": "77kg x 3",
    "Deadlift": "127kg x 3",
    "OHP": "47kg x 5",
}

# ---------------------------
# TEMPLATE FOR BLOCK 1
# ---------------------------
days = {
    "Day 1 (Lower Strength)": [
        ("Squat (Strength)", "3x3", starting_weights["Squat (Strength)"]),
        ("Deadlift", "3x3", starting_weights["Deadlift"]),
        ("Accessory Quad / Hamstring", "3x8–10", ""),
        ("Core", "3x12–15", ""),
    ],
    "Day 2 (Upper Strength)": [
        ("Bench Press (Strength)", "3x3", starting_weights["Bench (Strength)"]),
        ("Overhead Press", "3x5", starting_weights["OHP"]),
        ("Pull-ups / Rows", "3x8–12", ""),
        ("Barbell Shrugs", "3x10–12", ""),
    ],
    "Day 3 (Arms + Accessories)": [
        ("Barbell Curl", "3x8–12", ""),
        ("Skullcrusher", "3x8–12", ""),
        ("Dumbbell Hammer Curl", "3x10–12", ""),
        ("Triceps Pushdown", "3x10–12", ""),
    ],
    "Day 4 (Lower Hypertrophy)": [
        ("Front Squat / Hack Squat", "3x8–12", ""),
        ("Romanian Deadlift", "3x8–12", ""),
        ("Leg Press", "3x10–12", ""),
        ("Calves", "3x12–15", ""),
    ],
    "Day 5 (Upper Hypertrophy)": [
        ("Incline Bench Press", "3x8–12", ""),
        ("Dumbbell Overhead Press", "3x8–12", ""),
        ("Barbell Row", "3x8–12", ""),
        ("Dumbbell Shrugs", "3x12–15", ""),
    ],
    "Day 6 (Chest / Isolation)": [
        ("Cable Crossovers", "3x12–15", ""),
        ("Dumbbell Fly", "3x12–15", ""),
        ("Lateral Raises", "3x12–15", ""),
        ("Rear Delt Fly", "3x12–15", ""),
    ],
}

# ---------------------------
# BUILD EXCEL FILE
# ---------------------------
wb = Workbook()

# ---- Block 1 Sheet ----
ws = wb.active

if ws is None:
    ws = wb.create_sheet("Block 1")

ws.title = "Block 1"

headers = [
    "Day",
    "Exercise",
    "Sets/Reps",
    "Weight",
    "Progress Achieved (Y/N)",
    "Pain (0–10)",
    "Date",
]
ws.append(headers)

for day, exercises in days.items():
    for ex, reps, wgt in exercises:
        ws.append([day, ex, reps, wgt, "", "", ""])
    ws.append(["", "", "", "", "", "", ""])  # blank row between days

# ---- Weekly Summary Row ----
ws.append(["WEEKLY SUMMARY", "", "", "", '=COUNTIF(E2:E100, "Y")', "", ""])
last_row = ws.max_row

# ---- Block Summary Row ----
ws.append(["BLOCK SUMMARY", "", "", "", '=COUNTIF(E2:E100, "Y")', "", ""])

# ---------------------------
# FORMATTING
# ---------------------------
header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
for col in range(1, len(headers) + 1):
    ws.cell(row=1, column=col).fill = header_fill
    ws.cell(row=1, column=col).font = Font(bold=True)
    ws.cell(row=1, column=col).alignment = Alignment(horizontal="center")

# Conditional formatting for Pain column (red if ≥3)
pain_col = get_column_letter(headers.index("Pain (0–10)") + 1)
ws.conditional_formatting.add(
    f"{pain_col}2:{pain_col}{last_row}",
    CellIsRule(
        operator="greaterThanOrEqual",
        formula=["3"],
        stopIfTrue=True,
        fill=PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    ),
)

print("ws.columns", ws.columns)

# Adjust column widths
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2
    # print("col in columns: ", col)
    print("col[0].column: ", col[0].column)

# ---- Notes / Instructions Tab ----
notes = wb.create_sheet("Notes")
notes["A1"] = "Instructions"
notes["A1"].font = Font(bold=True)
notes["A2"] = (
    "1. Log weights manually each session.\n"
    "2. Progress Achieved = Y if top reps hit with good form (and Pain ≤ 2).\n"
    "3. No progression if Pain ≥ 3.\n"
    "4. Weekly Summary shows how many lifts progressed that week.\n"
    "5. Block Summary shows total progress across Block 1.\n"
    "6. All rep ranges are locked text for consistency.\n"
    "7. Use Notes tab for rehab tracking, tweaks, or extra comments."
)

# ---------------------------
# SAVE FILE
# ---------------------------
wb.save("C:\\Users\\Admin\\AppData\\Local\\Temp\\Block1_Training.xlsx")
print("✅ File Block1_Training.xlsx has been created successfully!")
