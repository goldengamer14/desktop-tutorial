import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.worksheet.table
import openpyxl.formatting.rule
from json import load

work_book = None
work_book = openpyxl.Workbook()
print("Created new Workbook")

work_sheet = work_book.active if work_book.active else work_book.create_sheet("Week 1")

exercise_routine = None
with open("Maaz-Python\\Openpyxl_Training\\Exercises.json", "r") as file:
    exercise_routine = load(file)
    file.close()

work_sheet.merge_cells("A1:F1")
work_sheet["A1"] = "Week 1"
work_sheet["A1"].alignment = openpyxl.styles.Alignment(
    horizontal="center", vertical="center"
)
work_sheet["A1"].font = openpyxl.styles.Font(bold=True, size=14, color="ececec")
work_sheet["A1"].fill = openpyxl.styles.PatternFill(
    start_color="4f4f4f", end_color="4f4f4f", fill_type="solid"
)
work_sheet.column_dimensions["A"].width = 30
work_sheet.row_dimensions[1].height = 20
work_sheet.column_dimensions["F"].width = 20
for col in ["B", "C", "D", "E"]:
    work_sheet.column_dimensions[col].width = 15

for day in exercise_routine:
    start_row = work_sheet.max_row + 3
    work_sheet.merge_cells(f"B{start_row}:F{start_row}")
    work_sheet[f"A{start_row}"] = day
    work_sheet[f"B{start_row}"] = f'{exercise_routine[day].get("type", "")}'
    work_sheet[f"A{start_row}"].font = work_sheet[f"B{start_row}"].font = (
        openpyxl.styles.Font(bold=True, size=12, color="ececec")
    )
    work_sheet[f"A{start_row}"].fill = work_sheet[f"B{start_row}"].fill = (
        openpyxl.styles.PatternFill(
            start_color="4f4f4f", end_color="4f4f4f", fill_type="solid"
        )
    )
    work_sheet[f"B{start_row}"].alignment = openpyxl.styles.Alignment(
        horizontal="center", vertical="center"
    )

    work_sheet.append(
        ["Exercise", "Sets", "Reps", "Reps Done", "Weight", "Pain Level(0-10)"]
    )
    table_start = f"A{work_sheet.max_row}"

    print(exercise_routine[day])
    for exercise, details in exercise_routine[day].items():
        print(exercise_routine[day][exercise])
        if exercise != "type":
            work_sheet.append(
                [
                    exercise,
                    details["Sets"],
                    details["Reps"],
                    "-",
                    details["Weight"],
                    "-",
                ]
            )
    table_end = (
        f"{openpyxl.utils.get_column_letter(work_sheet.max_column)}{work_sheet.max_row}"
    )

    print(f"Adding table for {day}: {table_start} to {table_end}")
    work_sheet.add_table(
        openpyxl.worksheet.table.Table(
            ref=f"{table_start}:{table_end}",
            displayName=f"Table_{str(day).replace(' ', '_')}",
        )
    )

    for header_row in work_sheet.iter_rows(
        min_row=work_sheet.max_row - len(exercise_routine[day]) + 1,
        max_row=work_sheet.max_row,
    ):
        for cell in header_row:
            cell.border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style="thick", color="000000"),
                right=openpyxl.styles.borders.Side(style="thick", color="000000"),
                top=openpyxl.styles.borders.Side(style="thick", color="000000"),
                bottom=openpyxl.styles.borders.Side(style="thick", color="000000"),
            )
            cell.font = openpyxl.styles.Font(bold=True, size=12, color="000000")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="ececec", end_color="ececec", fill_type="solid"
            )
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )

    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(
        f"{table_start}:{table_end}"
    )
    for row in work_sheet.iter_rows(
        min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row
    ):
        for cell in row:
            cell.border = openpyxl.styles.borders.Border(
                left=openpyxl.styles.borders.Side(style="thin", color="000000"),
                right=openpyxl.styles.borders.Side(style="thin", color="000000"),
                top=openpyxl.styles.borders.Side(style="thin", color="000000"),
                bottom=openpyxl.styles.borders.Side(style="thin", color="000000"),
            )
            cell.font = openpyxl.styles.Font(
                bold=False,
                size=11,
                color=(
                    "ececec" if not cell.row is None and cell.row % 2 == 0 else "000000"
                ),
            )
            cell.fill = (
                openpyxl.styles.PatternFill(
                    start_color="4f4f4f", end_color="4f4f4f", fill_type="solid"
                )
                if not cell.row is None and cell.row % 2 == 0
                else openpyxl.styles.PatternFill(
                    start_color="ffffff", end_color="ffffff", fill_type="solid"
                )
            )
            cell.alignment = openpyxl.styles.Alignment(
                horizontal="center", vertical="center"
            )

    red_fill = openpyxl.styles.PatternFill(
        start_color="af0000", end_color="af0000", fill_type="solid"
    )
    green_fill = openpyxl.styles.PatternFill(
        start_color="00af00", end_color="00af00", fill_type="solid"
    )
    yellow_fill = openpyxl.styles.PatternFill(
        start_color="afaf00", end_color="afaf00", fill_type="solid"
    )
    white_text = openpyxl.styles.Font(color="ececec")

    work_sheet.conditional_formatting.add(
        f"F{start_row+2}:F{work_sheet.max_row}",
        openpyxl.formatting.rule.CellIsRule(
            operator="between", formula=["0", "4"], fill=green_fill, font=white_text
        ),
    )
    work_sheet.conditional_formatting.add(
        f"F{start_row+2}:F{work_sheet.max_row}",
        openpyxl.formatting.rule.CellIsRule(
            operator="between", formula=["7", "10"], fill=red_fill, font=white_text
        ),
    )
    work_sheet.conditional_formatting.add(
        f"F{start_row+2}:F{work_sheet.max_row}",
        openpyxl.formatting.rule.CellIsRule(
            operator="between", formula=["4", "6"], fill=yellow_fill, font=white_text
        ),
    )

    for i in range(2):
        work_sheet.append([])

work_book.save("C:\\Users\\Admin\\AppData\\Local\\Temp\\Block1_Training.xlsx")
work_book.close()
